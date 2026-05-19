import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _base_style(ws, headers: list[str]):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def _set_column_widths(ws, widths: list[int]):
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def to_zoho_cc_xlsx(transactions: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "CreditCardStatement"

    headers = ["Date", "Purchases", "Payments/Refunds", "Payee", "Description", "Reference Number"]
    _base_style(ws, headers)

    for tx in transactions:
        purchases = tx["amount"] if not tx["is_credit"] else 0.0
        refunds = tx["amount"] if tx["is_credit"] else 0.0

        row = [
            tx["date"],
            purchases,
            refunds,
            tx["payee"],
            tx["description"],
            tx["reference"],
        ]
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = "DD/MM/YYYY"

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = "#,##0.00"

    _set_column_widths(ws, [14, 14, 18, 30, 60, 30])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_zoho_bank_xlsx(transactions: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "BankStatement"

    headers = ["Date", "Withdrawals", "Deposits", "Payee", "Description", "Reference Number"]
    _base_style(ws, headers)

    for tx in transactions:
        row = [
            tx["date"],
            tx["withdrawal"],
            tx["deposit"],
            tx["payee"],
            tx["description"],
            tx["reference"],
        ]
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = "DD/MM/YYYY"

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = "#,##0.00"

    _set_column_widths(ws, [14, 14, 14, 30, 70, 25])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def transactions_to_dataframe_cc(transactions: list[dict]) -> pd.DataFrame:
    rows = []
    for tx in transactions:
        rows.append({
            "Date": tx["date"].strftime("%d/%m/%Y"),
            "Purchases": tx["amount"] if not tx["is_credit"] else "",
            "Payments/Refunds": tx["amount"] if tx["is_credit"] else "",
            "Payee": tx["payee"],
            "Description": tx["description"],
            "Reference Number": tx["reference"],
        })
    return pd.DataFrame(rows)


def transactions_to_dataframe_bank(transactions: list[dict]) -> pd.DataFrame:
    rows = []
    for tx in transactions:
        rows.append({
            "Date": tx["date"].strftime("%d/%m/%Y"),
            "Withdrawals": tx["withdrawal"] if tx["withdrawal"] else "",
            "Deposits": tx["deposit"] if tx["deposit"] else "",
            "Payee": tx["payee"],
            "Description": tx["description"],
            "Reference Number": tx["reference"],
        })
    return pd.DataFrame(rows)
