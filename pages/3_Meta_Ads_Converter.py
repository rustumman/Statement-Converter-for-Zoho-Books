import io
from datetime import datetime

import openpyxl
import pandas as pd
import streamlit as st

# ── Pre-set Zoho fields ────────────────────────────────────────────────────────
PRESET = {
    "Bill Status": "Paid",
    "GST Treatment": "business_gst",
    "Is Inclusive Tax": True,
    "Vendor Name": "Meta",
    "Item Name": "Facebook Paid Campaign",
    "Account": "Digital Marketing Advertisement",
    "Quantity": 1,
    "Item Type": "service",
    "Tax Name": "IGST18",
    "Tax Percentage": 18,
}

ZOHO_COLUMNS = [
    "Bill Date",
    "Bill Number",
    "Rate",
    "Bill Status",
    "GST Treatment",
    "Is Inclusive Tax",
    "Vendor Name",
    "Item Name",
    "Account",
    "Quantity",
    "Item Type",
    "Tax Name",
    "Tax Percentage",
]


def parse_date(val) -> datetime | None:
    """Parse dates from Meta's invoice Excel.

    Meta's Excel file stores numeric dates with day and month swapped relative
    to what openpyxl reads (e.g. openpyxl yields 2026-01-03 but the real date
    is 3 Jan → 01 Mar, i.e. day=3, month=1 are transposed). We correct this by
    swapping month and day on datetime objects — matching the Excel formula:
        =DATE(YEAR(A2), DAY(A2), MONTH(A2))

    String dates arrive as DD/MM/YYYY and are parsed directly.
    """
    if isinstance(val, datetime):
        # Swap month ↔ day to undo Excel's misinterpretation
        try:
            return val.replace(month=val.day, day=val.month)
        except ValueError:
            return val  # swap not possible; return as-is
    if isinstance(val, str):
        val = val.strip()
        # Meta uses DD/MM/YYYY in string cells
        try:
            return datetime.strptime(val, "%d/%m/%Y")
        except ValueError:
            pass
        # Fallback formats
        for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
    return None


def load_workbook_from_csv(file_bytes: bytes) -> openpyxl.Workbook:
    """Convert a CSV file into an in-memory openpyxl Workbook so the same
    extract_transactions logic works for both CSV and Excel inputs."""
    import csv, io as _io
    text = file_bytes.decode("utf-8-sig", errors="replace")  # handle BOM
    reader = csv.reader(_io.StringIO(text))
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in reader:
        ws.append(row)
    return wb


def extract_transactions(wb: openpyxl.Workbook) -> pd.DataFrame:
    ws = wb.active

    # Find the header row and map column names → indices
    header_row_idx = None
    col_idx = {}
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if not row:
            continue
        row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
        if "date" in row_lower:
            header_row_idx = idx
            for i, name in enumerate(row_lower):
                col_idx[name] = i
            break

    if header_row_idx is None:
        raise ValueError("Could not find a 'Date' header row in the uploaded file.")

    # Resolve column positions by name (case-insensitive)
    date_col    = col_idx.get("date", 0)
    txnid_col   = col_idx.get("transaction id", 1)
    amount_col  = col_idx.get("amount", 3)   # explicit lookup by name

    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
        if not row or len(row) == 0:
            continue

        def cell(i):
            return row[i] if i < len(row) else None

        date_val = cell(date_col)
        txn_id   = cell(txnid_col)
        amount   = cell(amount_col)

        # Stop at footer (Total amount billed row has no date/txn_id)
        if date_val is None and txn_id is None:
            continue
        # Stop if any cell contains "total"
        if any("total" in str(c).lower() for c in row if c is not None):
            break
        if date_val is None:
            continue

        parsed_date = parse_date(date_val)
        if parsed_date is None:
            continue

        rows.append(
            {
                "Bill Date":   parsed_date,
                "Bill Number": str(txn_id).strip() if txn_id else "",
                "Rate":        amount,
            }
        )

    if not rows:
        raise ValueError("No transaction rows found after the header.")

    df = pd.DataFrame(rows)
    df["Bill Date"] = pd.to_datetime(df["Bill Date"])
    df["Rate"] = pd.to_numeric(df["Rate"], errors="coerce")
    for col, val in PRESET.items():
        df[col] = val

    return df[ZOHO_COLUMNS]


def to_xlsx(df: pd.DataFrame) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "in"

    ws.append(ZOHO_COLUMNS)

    for _, row in df.iterrows():
        ws.append(
            [
                row["Bill Date"].to_pydatetime() if hasattr(row["Bill Date"], "to_pydatetime") else row["Bill Date"],
                str(row["Bill Number"]),
                float(row["Rate"]) if pd.notna(row["Rate"]) else None,   # explicit Python float
                row["Bill Status"],
                row["GST Treatment"],
                bool(row["Is Inclusive Tax"]),
                row["Vendor Name"],
                row["Item Name"],
                row["Account"],
                int(row["Quantity"]),
                row["Item Type"],
                row["Tax Name"],
                float(row["Tax Percentage"]),
            ]
        )

    # Format Bill Date column as DD/MM/YYYY
    date_fmt = openpyxl.styles.numbers.FORMAT_DATE_DMYSLASH
    for cell in ws["A"][1:]:  # skip header
        cell.number_format = date_fmt

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── UI ─────────────────────────────────────────────────────────────────────────
st.title("Meta Ads → Zoho Books Converter")
st.markdown(
    "Upload a Meta Ads invoice summary (.xlsx) and download a Zoho Books "
    "bulk-import Excel file ready to use."
)

uploaded = st.file_uploader(
    "Upload Meta Ads invoice summary (.xlsx or .csv)",
    type=["xlsx", "csv"],
    accept_multiple_files=False,
)

if not uploaded:
    st.info("Upload a file to get started.")
    st.stop()

try:
    file_bytes = uploaded.read()
    if uploaded.name.lower().endswith(".csv"):
        wb = load_workbook_from_csv(file_bytes)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    df = extract_transactions(wb)
except Exception as e:
    st.error(f"Could not parse file: {e}")
    st.stop()

st.success(f"Parsed **{len(df)} transactions**.")

# Date range summary
min_date = df["Bill Date"].min().strftime("%d %b %Y")
max_date = df["Bill Date"].max().strftime("%d %b %Y")
total_amount = df["Rate"].sum()
st.markdown(
    f"**Period:** {min_date} – {max_date} &nbsp;|&nbsp; "
    f"**Total billed:** ₹{total_amount:,.2f}"
)

st.dataframe(
    df.assign(**{"Bill Date": df["Bill Date"].dt.strftime("%d/%m/%Y")}),
    use_container_width=True,
    hide_index=True,
)

xlsx_bytes = to_xlsx(df)
filename = f"meta_ads_zoho_{df['Bill Date'].min().strftime('%Y%m%d')}_{df['Bill Date'].max().strftime('%Y%m%d')}.xlsx"

st.download_button(
    label="⬇ Download Zoho Import Excel",
    data=xlsx_bytes,
    file_name=filename,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
